from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, send_file
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta
from PIL import Image
import os
import sqlite3
import base64
from functools import wraps
import io
import pytz

# Set timezone to IST
IST = pytz.timezone('Asia/Kolkata')

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-change-in-production'
app.config['UPLOAD_FOLDER'] = 'static/uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Ensure upload directory exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Flask-Login setup
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Database setup
def get_db():
    conn = sqlite3.connect('attendance.db')
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    conn.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            email TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'user',
            location_enabled INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS attendance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            check_in_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            check_out_time TIMESTAMP,
            front_image_path TEXT,
            rear_image_path TEXT,
            checkout_front_image_path TEXT,
            checkout_rear_image_path TEXT,
            additional_image_1 TEXT,
            additional_image_2 TEXT,
            additional_image_3 TEXT,
            additional_image_4 TEXT,
            checkin_latitude REAL,
            checkin_longitude REAL,
            checkout_latitude REAL,
            checkout_longitude REAL,
            city TEXT,
            full_address TEXT,
            checkout_city TEXT,
            checkout_full_address TEXT,
            status TEXT DEFAULT 'checked_in',
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
    ''')
    
    # Add additional image columns if they don't exist (migration)
    try:
        cursor = conn.cursor()
        cursor.execute("PRAGMA table_info(attendance)")
        columns = [column[1] for column in cursor.fetchall()]
        
        if 'additional_image_1' not in columns:
            conn.execute('ALTER TABLE attendance ADD COLUMN additional_image_1 TEXT')
        if 'additional_image_2' not in columns:
            conn.execute('ALTER TABLE attendance ADD COLUMN additional_image_2 TEXT')
        if 'additional_image_3' not in columns:
            conn.execute('ALTER TABLE attendance ADD COLUMN additional_image_3 TEXT')
        if 'additional_image_4' not in columns:
            conn.execute('ALTER TABLE attendance ADD COLUMN additional_image_4 TEXT')
        
        conn.commit()
    except Exception as e:
        print(f"Migration error: {e}")
    
    # Create default admin if not exists
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM users WHERE username = 'admin'")
    if not cursor.fetchone():
        admin_hash = generate_password_hash('admin@vsconstruction2025!')
        cursor.execute(
            "INSERT INTO users (username, email, password_hash, role) VALUES (?, ?, ?, ?)",
            ('admin', 'admin@vscattendance.com', admin_hash, 'admin')
        )
        conn.commit()
    conn.close()

# User class for Flask-Login
class User(UserMixin):
    def __init__(self, id, username, email, role, location_enabled=1):
        self.id = id
        self.username = username
        self.email = email
        self.role = role
        self.location_enabled = location_enabled

@login_manager.user_loader
def load_user(user_id):
    conn = get_db()
    user = conn.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
    conn.close()

    if user:
        return User(
            user['id'],
            user['username'],
            user['email'],
            user['role'],
            user['location_enabled']
        )
    return None


# Admin required decorator
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'admin':
            flash('Admin access required', 'danger')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Routes
@app.route('/')
def index():
    if current_user.is_authenticated:
        if current_user.role == 'admin':
            return redirect(url_for('admin_dashboard'))
        return redirect(url_for('user_dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        conn = get_db()
        user = conn.execute('SELECT * FROM users WHERE username = ?', (username,)).fetchone()
        conn.close()
        
        if user and check_password_hash(user['password_hash'], password):
            user_obj = User(user['id'], user['username'], user['email'], user['role'])
            login_user(user_obj)
            flash('Login successful!', 'success')
            return redirect(url_for('index'))
        else:
            flash('Invalid username or password', 'danger')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Logged out successfully', 'success')
    return redirect(url_for('login'))

@app.route('/admin/dashboard')
@login_required
@admin_required
def admin_dashboard():
    conn = get_db()
    users = conn.execute('SELECT id, username, email, role, location_enabled, created_at FROM users').fetchall()
    attendance = conn.execute('''
        SELECT a.*, u.username 
        FROM attendance a 
        JOIN users u ON a.user_id = u.id 
        ORDER BY a.check_in_time DESC 
        LIMIT 50
    ''').fetchall()
    conn.close()
    return render_template('admin_dashboard.html', users=users, attendance=attendance)

@app.route('/admin/add_user', methods=['POST'])
@login_required
@admin_required
def add_user():
    username = request.form.get('username')
    email = request.form.get('email')
    password = request.form.get('password')
    role = request.form.get('role', 'user')
    location_enabled = 1 if request.form.get('location_enabled') == 'on' else 0
    
    if not username or not email or not password:
        flash('All fields are required', 'danger')
        return redirect(url_for('admin_dashboard'))
    
    conn = get_db()
    try:
        password_hash = generate_password_hash(password)
        conn.execute(
            'INSERT INTO users (username, email, password_hash, role, location_enabled) VALUES (?, ?, ?, ?, ?)',
            (username, email, password_hash, role, location_enabled)
        )
        conn.commit()
        location_status = "with" if location_enabled else "without"
        flash(f'User {username} added successfully {location_status} location tracking', 'success')
    except sqlite3.IntegrityError:
        flash('Username or email already exists', 'danger')
    finally:
        conn.close()
    
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/toggle_location/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def toggle_location(user_id):
    conn = get_db()
    user = conn.execute('SELECT * FROM users WHERE id = ?', (user_id,)).fetchone()
    
    if user:
        new_status = 0 if user['location_enabled'] else 1
        conn.execute('UPDATE users SET location_enabled = ? WHERE id = ?', (new_status, user_id))
        conn.commit()
        status_text = "enabled" if new_status else "disabled"
        flash(f'Location tracking {status_text} for {user["username"]}', 'success')
    else:
        flash('User not found', 'danger')
    
    conn.close()
    return redirect(url_for('admin_dashboard'))

@app.route('/user/dashboard')
@login_required
def user_dashboard():
    conn = get_db()
    attendance = conn.execute(
        'SELECT * FROM attendance WHERE user_id = ? ORDER BY check_in_time DESC LIMIT 10',
        (current_user.id,)
    ).fetchall()
    
    # Check if user is currently checked in
    current_checkin = conn.execute(
        'SELECT * FROM attendance WHERE user_id = ? AND status = "checked_in" ORDER BY check_in_time DESC LIMIT 1',
        (current_user.id,)
    ).fetchone()
    conn.close()
    
    return render_template('user_dashboard.html', attendance=attendance, current_checkin=current_checkin)

@app.route('/user/checkin')
@login_required
def checkin_page():
    # Check if already checked in
    conn = get_db()
    current_checkin = conn.execute(
        'SELECT * FROM attendance WHERE user_id = ? AND status = "checked_in"',
        (current_user.id,)
    ).fetchone()
    conn.close()
    
    if current_checkin:
        flash('You are already checked in. Please check out first.', 'warning')
        return redirect(url_for('user_dashboard'))
    
    # Pass location permission to template
    return render_template('checkin.html', location_allowed=current_user.location_enabled)

@app.route('/user/checkout')
@login_required
def checkout_page():
    # Check if user is checked in
    conn = get_db()
    current_checkin = conn.execute(
        'SELECT * FROM attendance WHERE user_id = ? AND status = "checked_in"',
        (current_user.id,)
    ).fetchone()
    conn.close()
    
    if not current_checkin:
        flash('You need to check in first.', 'warning')
        return redirect(url_for('user_dashboard'))
    
    # Pass location permission to template
    return render_template('checkout.html', location_allowed=current_user.location_enabled)

@app.route('/api/checkin', methods=['POST'])
@login_required
def api_checkin():
    data = request.json
    
    # Save images
    front_image = data.get('front_image')
    rear_image = data.get('rear_image')
    latitude = data.get('latitude')
    longitude = data.get('longitude')
    city = data.get('city', 'Location not captured')
    full_address = data.get('full_address', 'Location not captured')
    
    # Use IST timezone
    timestamp = datetime.now(IST).strftime('%Y%m%d_%H%M%S')
    checkin_time = datetime.now(IST).strftime('%Y-%m-%d %H:%M:%S')
    
    front_filename = f"checkin_front_{current_user.id}_{timestamp}.jpg"
    rear_filename = f"checkin_rear_{current_user.id}_{timestamp}.jpg"
    
    # Save and create thumbnails for front image
    if front_image:
        front_image_data = front_image.split(',')[1]
        front_path = os.path.join(app.config['UPLOAD_FOLDER'], front_filename)
        with open(front_path, 'wb') as f:
            f.write(base64.b64decode(front_image_data))
        
        # Create thumbnail
        create_thumbnail(front_path, f"thumb_{front_filename}")
    
    # Save and create thumbnails for rear image
    if rear_image:
        rear_image_data = rear_image.split(',')[1]
        rear_path = os.path.join(app.config['UPLOAD_FOLDER'], rear_filename)
        with open(rear_path, 'wb') as f:
            f.write(base64.b64decode(rear_image_data))
        
        # Create thumbnail
        create_thumbnail(rear_path, f"thumb_{rear_filename}")
    
    # Save to database
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO attendance (user_id, front_image_path, rear_image_path, checkin_latitude, checkin_longitude, city, full_address, check_in_time)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    ''', (current_user.id, front_filename, rear_filename, latitude, longitude, city, full_address, checkin_time))
    attendance_id = cursor.lastrowid
    conn.commit()
    conn.close()
    
    return jsonify({'success': True, 'message': 'Check-in successful!', 'attendance_id': attendance_id})

@app.route('/api/checkout', methods=['POST'])
@login_required
def api_checkout():
    data = request.json
    
    # Get checkout images and location
    checkout_front_image = data.get('checkout_front_image')
    checkout_rear_image = data.get('checkout_rear_image')
    checkout_latitude = data.get('checkout_latitude')
    checkout_longitude = data.get('checkout_longitude')
    checkout_city = data.get('checkout_city', 'Location not captured')
    checkout_full_address = data.get('checkout_full_address', 'Location not captured')
    
    # Use IST timezone
    timestamp = datetime.now(IST).strftime('%Y%m%d_%H%M%S')
    checkout_time = datetime.now(IST).strftime('%Y-%m-%d %H:%M:%S')
    checkout_front_filename = None
    checkout_rear_filename = None
    
    # Save checkout front image
    if checkout_front_image:
        checkout_front_filename = f"checkout_front_{current_user.id}_{timestamp}.jpg"
        checkout_front_image_data = checkout_front_image.split(',')[1]
        checkout_front_path = os.path.join(app.config['UPLOAD_FOLDER'], checkout_front_filename)
        with open(checkout_front_path, 'wb') as f:
            f.write(base64.b64decode(checkout_front_image_data))
        create_thumbnail(checkout_front_path, f"thumb_{checkout_front_filename}")
    
    # Save checkout rear image
    if checkout_rear_image:
        checkout_rear_filename = f"checkout_rear_{current_user.id}_{timestamp}.jpg"
        checkout_rear_image_data = checkout_rear_image.split(',')[1]
        checkout_rear_path = os.path.join(app.config['UPLOAD_FOLDER'], checkout_rear_filename)
        with open(checkout_rear_path, 'wb') as f:
            f.write(base64.b64decode(checkout_rear_image_data))
        create_thumbnail(checkout_rear_path, f"thumb_{checkout_rear_filename}")
    
    conn = get_db()
    conn.execute('''
        UPDATE attendance 
        SET check_out_time = ?, 
            status = 'checked_out',
            checkout_front_image_path = ?,
            checkout_rear_image_path = ?,
            checkout_latitude = ?,
            checkout_longitude = ?,
            checkout_city = ?,
            checkout_full_address = ?
        WHERE user_id = ? AND status = 'checked_in'
    ''', (checkout_time, checkout_front_filename, checkout_rear_filename, checkout_latitude, checkout_longitude, 
          checkout_city, checkout_full_address, current_user.id))
    conn.commit()
    conn.close()
    
    return jsonify({'success': True, 'message': 'Check-out successful!'})

@app.route('/api/add-images/<int:attendance_id>', methods=['POST'])
@login_required
def add_additional_images(attendance_id):
    """Add additional images to an existing check-in"""
    try:
        data = request.get_json()
        
        if not data or 'additional_images' not in data:
            return jsonify({'success': False, 'error': 'No images provided'}), 400
        
        # Find the attendance record
        conn = get_db()
        attendance = conn.execute(
            'SELECT * FROM attendance WHERE id = ? AND user_id = ? AND status = ?',
            (attendance_id, current_user.id, 'checked_in')
        ).fetchone()
        
        if not attendance:
            conn.close()
            return jsonify({'success': False, 'error': 'Active check-in not found'}), 404
        
        # Check how many additional images already exist
        existing_count = sum([
            1 if attendance['additional_image_1'] else 0,
            1 if attendance['additional_image_2'] else 0,
            1 if attendance['additional_image_3'] else 0,
            1 if attendance['additional_image_4'] else 0
        ])
        
        if existing_count >= 4:
            conn.close()
            return jsonify({'success': False, 'error': 'Maximum 4 additional images already uploaded'}), 400
        
        # Process and save new images
        additional_images = data['additional_images']
        uploaded_files = []
        timestamp = datetime.now(IST).strftime('%Y%m%d_%H%M%S_%f')
        
        # Determine which slots to fill
        slots_to_fill = []
        if not attendance['additional_image_1']:
            slots_to_fill.append('additional_image_1')
        if not attendance['additional_image_2']:
            slots_to_fill.append('additional_image_2')
        if not attendance['additional_image_3']:
            slots_to_fill.append('additional_image_3')
        if not attendance['additional_image_4']:
            slots_to_fill.append('additional_image_4')
        
        # Save images to the available slots
        for idx, image_data in enumerate(additional_images[:len(slots_to_fill)]):
            filename = f"additional_{current_user.id}_{attendance_id}_{timestamp}_{idx}.jpg"
            
            # Save image
            image_data_clean = image_data.split(',')[1]
            image_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            with open(image_path, 'wb') as f:
                f.write(base64.b64decode(image_data_clean))
            
            # Create thumbnail
            create_thumbnail(image_path, f"thumb_{filename}")
            
            uploaded_files.append((slots_to_fill[idx], filename))
        
        # Update database with new image paths
        if uploaded_files:
            update_parts = []
            update_values = []
            for slot, filename in uploaded_files:
                update_parts.append(f"{slot} = ?")
                update_values.append(filename)
            
            update_values.append(attendance_id)
            conn.execute(
                f"UPDATE attendance SET {', '.join(update_parts)} WHERE id = ?",
                update_values
            )
            conn.commit()
        
        conn.close()
        
        return jsonify({
            'success': True,
            'message': f'{len(uploaded_files)} image(s) uploaded successfully',
            'images_uploaded': len(uploaded_files),
            'total_additional_images': existing_count + len(uploaded_files)
        })
        
    except Exception as e:
        print(f"Error adding additional images: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

# Helper function to create thumbnails
def create_thumbnail(image_path, thumb_filename, size=(150, 150)):
    try:
        with Image.open(image_path) as img:
            img.thumbnail(size, Image.Resampling.LANCZOS)
            thumb_path = os.path.join(app.config['UPLOAD_FOLDER'], thumb_filename)
            img.save(thumb_path, 'JPEG', quality=85)
    except Exception as e:
        print(f"Error creating thumbnail: {e}")

# Admin routes
@app.route('/admin/delete_all_records', methods=['POST'])
@login_required
@admin_required
def delete_all_records():
    admin_password = request.form.get('admin_password')
    
    # Verify admin password
    conn = get_db()
    admin = conn.execute('SELECT * FROM users WHERE id = ?', (current_user.id,)).fetchone()
    
    if not admin or not check_password_hash(admin['password_hash'], admin_password):
        conn.close()
        flash('Invalid password. Delete operation cancelled.', 'danger')
        return redirect(url_for('admin_dashboard'))
    
    try:
        # Delete all attendance records
        conn.execute('DELETE FROM attendance')
        conn.commit()
        conn.close()
        
        # Delete all uploaded images
        upload_folder = app.config['UPLOAD_FOLDER']
        for filename in os.listdir(upload_folder):
            if filename.endswith(('.jpg', '.jpeg', '.png')):
                file_path = os.path.join(upload_folder, filename)
                try:
                    os.remove(file_path)
                except:
                    pass
        
        flash('All attendance records deleted successfully', 'success')
    except Exception as e:
        flash(f'Error deleting records: {str(e)}', 'danger')
    
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/export_report')
@login_required
@admin_required
def export_report():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    if not start_date or not end_date:
        flash('Please provide start and end dates', 'danger')
        return redirect(url_for('admin_dashboard'))
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Report"
        
        # Headers
        headers = ['User', 'Check In', 'Check Out', 'Duration (hrs)', 'Check-in Location', 
                   'Check-out Location', 'Additional Images', 'Status']
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Get data with date filter
        conn = get_db()
        records = conn.execute('''
            SELECT u.username, a.check_in_time, a.check_out_time, a.city, 
                   a.checkout_city, a.status, a.checkin_latitude, a.checkin_longitude,
                   a.checkout_latitude, a.checkout_longitude, a.full_address, a.checkout_full_address,
                   a.additional_image_1, a.additional_image_2, a.additional_image_3, a.additional_image_4
            FROM attendance a
            JOIN users u ON a.user_id = u.id
            WHERE DATE(a.check_in_time) BETWEEN ? AND ?
            ORDER BY a.check_in_time DESC
        ''', (start_date, end_date)).fetchall()
        conn.close()
        
        # Add data
        for record in records:
            duration = 'N/A'
            if record['check_out_time']:
                try:
                    checkin = datetime.strptime(record['check_in_time'], '%Y-%m-%d %H:%M:%S')
                    checkout = datetime.strptime(record['check_out_time'], '%Y-%m-%d %H:%M:%S')
                    duration_hours = (checkout - checkin).total_seconds() / 3600
                    duration = f"{duration_hours:.2f}"
                except:
                    pass
            
            checkin_loc = record['city'] or 'Not captured'
            if record['full_address'] and record['full_address'] != record['city']:
                checkin_loc = record['full_address']
            if record['checkin_latitude'] and record['checkin_longitude'] and record['checkin_latitude'] != 0:
                checkin_loc += f" ({record['checkin_latitude']:.4f}, {record['checkin_longitude']:.4f})"
            
            checkout_loc = record['checkout_city'] or 'Not captured'
            if record['checkout_full_address'] and record['checkout_full_address'] != record['checkout_city']:
                checkout_loc = record['checkout_full_address']
            if record['checkout_latitude'] and record['checkout_longitude'] and record['checkout_latitude'] != 0:
                checkout_loc += f" ({record['checkout_latitude']:.4f}, {record['checkout_longitude']:.4f})"
            
            # Count additional images
            additional_count = sum([
                1 if record['additional_image_1'] else 0,
                1 if record['additional_image_2'] else 0,
                1 if record['additional_image_3'] else 0,
                1 if record['additional_image_4'] else 0
            ])
            
            ws.append([
                record['username'],
                record['check_in_time'],
                record['check_out_time'] or 'N/A',
                duration,
                checkin_loc,
                checkout_loc,
                f"{additional_count}/4",
                record['status']
            ])
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f'attendance_report_{start_date}_to_{end_date}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except ImportError:
        flash('openpyxl library not installed. Run: pip install openpyxl', 'danger')
        return redirect(url_for('admin_dashboard'))
    except Exception as e:
        flash(f'Error generating report: {str(e)}', 'danger')
        return redirect(url_for('admin_dashboard'))

if __name__ == '__main__':
    init_db()
    app.run(debug=True, host='0.0.0.0', port=5000)